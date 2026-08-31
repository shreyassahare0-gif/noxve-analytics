"""
Builds NOXVE_Sales_Report.xlsx:
 - Raw_Orders / Raw_Order_Items / Products  -> raw data tables
 - Monthly_Summary, Product_Performance, Payment_Method_Analysis,
   Customer_Summary, Dashboard_Summary -> ALL FORMULA-DRIVEN (SUMIFS/COUNTIFS/
   AVERAGEIFS/VLOOKUP/IF), nothing hardcoded, so it recalculates if raw data changes.
 - Instructions sheet with practice exercises tied to the Advanced Excel course module.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE, "..", "data")
OUT_PATH = os.path.join(BASE, "..", "excel", "NOXVE_Sales_Report.xlsx")
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)

orders = pd.read_csv(os.path.join(DATA_DIR, "orders.csv"), parse_dates=["order_datetime"])
items = pd.read_csv(os.path.join(DATA_DIR, "order_items.csv"))
products = pd.read_csv(os.path.join(DATA_DIR, "products.csv"))
customers = pd.read_csv(os.path.join(DATA_DIR, "customers.csv"))

orders["order_date"] = orders["order_datetime"].dt.date.astype(str)
orders["order_month"] = orders["order_datetime"].dt.to_period("M").astype(str)

items_full = items.merge(products, on="product_id").merge(
    orders[["order_id", "order_date", "order_month", "order_status"]], on="order_id"
)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="111111")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="B08D57")
LABEL_FONT = Font(name=FONT_NAME, bold=True)
NORMAL_FONT = Font(name=FONT_NAME)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
wb.remove(wb.active)


def write_table(ws, df, start_row=1, header_fill=True):
    """Write a pandas dataframe as a raw data table with a header row."""
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=col)
        c.font = HEADER_FONT
        if header_fill:
            c.fill = HEADER_FILL
        c.border = BORDER
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = NORMAL_FONT
            c.border = BORDER
    for j, col in enumerate(df.columns, start=1):
        width = max(12, min(28, int(df[col].astype(str).str.len().mean()) + 6))
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = "A2"
    return start_row + 1, start_row + len(df)  # first data row, last data row


# ---------------------------------------------------------------------------
# SHEET: Products (reference / lookup table)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Products")
first_p, last_p = write_table(ws, products[["product_id", "product_name", "collection",
                                             "category", "price", "unit_cost"]])
# VLOOKUP will search by product_name, so the range must start at column B (product_name)
PROD_RANGE_NAME = f"Products!$B${first_p}:$F${last_p}"

# ---------------------------------------------------------------------------
# SHEET: Raw_Orders
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Raw_Orders")
raw_orders_cols = ["order_id", "order_date", "order_month", "customer_id", "payment_method",
                    "order_status", "discount_code", "gross_amount", "discount_amount",
                    "shipping_fee", "net_revenue"]
first_o, last_o = write_table(ws, orders[raw_orders_cols])
OD = {c: get_column_letter(i + 1) for i, c in enumerate(raw_orders_cols)}

# ---------------------------------------------------------------------------
# SHEET: Raw_Order_Items
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Raw_Order_Items")
raw_items_cols = ["order_item_id", "order_id", "order_date", "order_month", "product_id",
                   "product_name", "category", "quantity", "unit_price", "line_total", "order_status"]
first_i, last_i = write_table(ws, items_full[raw_items_cols])
ID = {c: get_column_letter(i + 1) for i, c in enumerate(raw_items_cols)}

# ---------------------------------------------------------------------------
# SHEET: Monthly_Summary  (COUNTIFS / SUMIFS / formula-driven)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Monthly_Summary")
ws["A1"] = "NOXVE — Monthly Performance (Delivered Orders Only)"
ws["A1"].font = TITLE_FONT
headers = ["Month", "Delivered Orders", "Net Revenue (Rs.)", "Gross Revenue (Rs.)",
           "Discount Given (Rs.)", "Avg Order Value (Rs.)"]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

months = sorted(orders["order_month"].unique())
for i, m in enumerate(months, start=4):
    ws.cell(row=i, column=1, value=m).border = BORDER
    rng_month = f"Raw_Orders!${OD['order_month']}${first_o}:${OD['order_month']}${last_o}"
    rng_status = f"Raw_Orders!${OD['order_status']}${first_o}:${OD['order_status']}${last_o}"
    rng_net = f"Raw_Orders!${OD['net_revenue']}${first_o}:${OD['net_revenue']}${last_o}"
    rng_gross = f"Raw_Orders!${OD['gross_amount']}${first_o}:${OD['gross_amount']}${last_o}"
    rng_disc = f"Raw_Orders!${OD['discount_amount']}${first_o}:${OD['discount_amount']}${last_o}"

    ws.cell(row=i, column=2, value=f'=COUNTIFS({rng_month},A{i},{rng_status},"Delivered")').border = BORDER
    ws.cell(row=i, column=3, value=f'=SUMIFS({rng_net},{rng_month},A{i},{rng_status},"Delivered")').border = BORDER
    ws.cell(row=i, column=4, value=f'=SUMIFS({rng_gross},{rng_month},A{i},{rng_status},"Delivered")').border = BORDER
    ws.cell(row=i, column=5, value=f'=SUMIFS({rng_disc},{rng_month},A{i},{rng_status},"Delivered")').border = BORDER
    ws.cell(row=i, column=6, value=f'=IFERROR(C{i}/B{i},0)').border = BORDER

last_month_row = 3 + len(months)
ws.cell(row=last_month_row + 1, column=1, value="TOTAL").font = LABEL_FONT
for col_letter in ["B", "C", "D", "E"]:
    ws.cell(row=last_month_row + 1, column=ord(col_letter) - 64,
            value=f"=SUM({col_letter}4:{col_letter}{last_month_row})").font = LABEL_FONT
for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].width = 20
ws.freeze_panes = "A4"

# ---------------------------------------------------------------------------
# SHEET: Product_Performance (SUMIFS + VLOOKUP)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Product_Performance")
ws["A1"] = "NOXVE — Product Performance (Delivered Orders Only)"
ws["A1"].font = TITLE_FONT
headers = ["Product Name", "Category (VLOOKUP)", "Price (VLOOKUP)", "Unit Cost (VLOOKUP)",
           "Units Sold", "Revenue (Rs.)", "Est. Profit (Rs.)"]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

prod_names = products["product_name"].tolist()
for i, pname in enumerate(prod_names, start=4):
    ws.cell(row=i, column=1, value=pname).border = BORDER
    # Range starts at product_name (col B of Products), so: col2=category, col3=price, col4=unit_cost
    ws.cell(row=i, column=2, value=f'=VLOOKUP(A{i},{PROD_RANGE_NAME},3,FALSE)').border = BORDER
    ws.cell(row=i, column=3, value=f'=VLOOKUP(A{i},{PROD_RANGE_NAME},4,FALSE)').border = BORDER
    ws.cell(row=i, column=4, value=f'=VLOOKUP(A{i},{PROD_RANGE_NAME},5,FALSE)').border = BORDER

    rng_name = f"Raw_Order_Items!${ID['product_name']}${first_i}:${ID['product_name']}${last_i}"
    rng_status2 = f"Raw_Order_Items!${ID['order_status']}${first_i}:${ID['order_status']}${last_i}"
    rng_qty = f"Raw_Order_Items!${ID['quantity']}${first_i}:${ID['quantity']}${last_i}"
    rng_lt = f"Raw_Order_Items!${ID['line_total']}${first_i}:${ID['line_total']}${last_i}"

    ws.cell(row=i, column=5, value=f'=SUMIFS({rng_qty},{rng_name},A{i},{rng_status2},"Delivered")').border = BORDER
    ws.cell(row=i, column=6, value=f'=SUMIFS({rng_lt},{rng_name},A{i},{rng_status2},"Delivered")').border = BORDER
    ws.cell(row=i, column=7, value=f'=F{i}-(E{i}*D{i})').border = BORDER

last_prod_row = 3 + len(prod_names)
ws.cell(row=last_prod_row + 1, column=1, value="TOTAL").font = LABEL_FONT
for col_letter in ["E", "F", "G"]:
    ws.cell(row=last_prod_row + 1, column=ord(col_letter) - 64,
            value=f"=SUM({col_letter}4:{col_letter}{last_prod_row})").font = LABEL_FONT
for col in range(1, 8):
    ws.column_dimensions[get_column_letter(col)].width = 20
ws.freeze_panes = "A4"

# ---------------------------------------------------------------------------
# SHEET: Payment_Method_Analysis (COD vs Prepaid — COUNTIFS/AVERAGEIFS)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Payment_Method_Analysis")
ws["A1"] = "NOXVE — COD vs Prepaid Performance"
ws["A1"].font = TITLE_FONT
headers = ["Payment Method", "Total Orders", "RTO Orders", "RTO Rate %",
           "Delivered Orders", "Avg Order Value (Delivered, Rs.)"]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

rng_pay = f"Raw_Orders!${OD['payment_method']}${first_o}:${OD['payment_method']}${last_o}"
rng_status = f"Raw_Orders!${OD['order_status']}${first_o}:${OD['order_status']}${last_o}"
rng_net = f"Raw_Orders!${OD['net_revenue']}${first_o}:${OD['net_revenue']}${last_o}"

for i, pm in enumerate(["COD", "Prepaid"], start=4):
    ws.cell(row=i, column=1, value=pm).border = BORDER
    ws.cell(row=i, column=2, value=f'=COUNTIFS({rng_pay},A{i})').border = BORDER
    ws.cell(row=i, column=3, value=f'=COUNTIFS({rng_pay},A{i},{rng_status},"RTO")').border = BORDER
    ws.cell(row=i, column=4, value=f'=IFERROR(C{i}/B{i},0)').border = BORDER
    ws.cell(row=i, column=4).number_format = "0.0%"
    ws.cell(row=i, column=5, value=f'=COUNTIFS({rng_pay},A{i},{rng_status},"Delivered")').border = BORDER
    ws.cell(row=i, column=6, value=f'=IFERROR(AVERAGEIFS({rng_net},{rng_pay},A{i},{rng_status},"Delivered"),0)').border = BORDER
for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].width = 22

# ---------------------------------------------------------------------------
# SHEET: Customer_Summary (per-customer COUNTIFS/SUMIFS + New vs Repeat flag)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Customer_Summary")
unique_customers = sorted(orders["customer_id"].unique())
headers = ["customer_id", "Order Count (Delivered)", "Total Spent (Rs.)", "Customer Type"]
for j, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.border = BORDER

rng_cust = f"Raw_Orders!${OD['customer_id']}${first_o}:${OD['customer_id']}${last_o}"
for i, cid in enumerate(unique_customers, start=2):
    ws.cell(row=i, column=1, value=cid).border = BORDER
    ws.cell(row=i, column=2, value=f'=COUNTIFS({rng_cust},A{i},{rng_status},"Delivered")').border = BORDER
    ws.cell(row=i, column=3, value=f'=SUMIFS({rng_net},{rng_cust},A{i},{rng_status},"Delivered")').border = BORDER
    ws.cell(row=i, column=4, value=f'=IF(B{i}>1,"Repeat","New")').border = BORDER
last_cust_row = 1 + len(unique_customers)
ws.freeze_panes = "A2"
for col in range(1, 5):
    ws.column_dimensions[get_column_letter(col)].width = 20

# ---------------------------------------------------------------------------
# SHEET: Dashboard_Summary (top-level KPIs, formulas referencing other sheets)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Dashboard_Summary", 0)
ws["A1"] = "NOXVE — Year 1 Performance Dashboard"
ws["A1"].font = Font(name=FONT_NAME, size=16, bold=True, color="B08D57")
ws["A2"] = "Synthetic dataset for portfolio purposes — modeled on NOXVE's real catalog & operations"
ws["A2"].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")

kpis = [
    ("Total Delivered Orders", f"=Monthly_Summary!B{last_month_row+1}"),
    ("Total Net Revenue (Rs.)", f"=Monthly_Summary!C{last_month_row+1}"),
    ("Overall Avg Order Value (Rs.)", "=B5/B4"),
    ("Total Customers", f"=COUNTA(Customer_Summary!A2:A{last_cust_row})"),
    ("Repeat Customers", f'=COUNTIF(Customer_Summary!D2:D{last_cust_row},"Repeat")'),
    ("Repeat Customer Rate %", "=B8/B7"),
    ("COD RTO Rate %", "=Payment_Method_Analysis!D4"),
    ("Prepaid RTO Rate %", "=Payment_Method_Analysis!D5"),
]
row = 4
for label, formula in kpis:
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    fc = ws.cell(row=row, column=2, value=formula)
    if "Rate" in label:
        fc.number_format = "0.0%"
    elif "Revenue" in label or "Value" in label:
        fc.number_format = "Rs. #,##0"
    row += 1
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 20

# ---------------------------------------------------------------------------
# SHEET: Instructions (project story + practice exercises)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Instructions")
ws["A1"] = "How to use this workbook"
ws["A1"].font = TITLE_FONT
lines = [
    "",
    "This workbook is built for practicing the Advanced Excel for Data Analytics module.",
    "All summary sheets (Monthly_Summary, Product_Performance, Payment_Method_Analysis,",
    "Customer_Summary, Dashboard_Summary) use live formulas — SUMIFS, COUNTIFS, AVERAGEIFS,",
    "VLOOKUP, IF, IFERROR — referencing the raw data sheets. Nothing is hardcoded.",
    "",
    "Try it yourself (maps to your Advanced Excel lessons):",
    "1. Insert a PivotTable from Raw_Orders to reproduce Monthly_Summary from scratch.",
    "2. Add conditional formatting to Payment_Method_Analysis!D4:D5 (RTO Rate) — red if >10%.",
    "3. Build a PivotChart: Revenue by Month, split by Payment Method.",
    "4. Add a helper column in Raw_Orders using TEXT()/LEFT() to extract order_month",
    "   yourself instead of reading the pre-built column, then rebuild Monthly_Summary.",
    "5. In Customer_Summary, add a 5th column ranking customers by Total Spent using",
    "   RANK() or LARGE(), and pull the Top 10 into a new sheet with INDEX/MATCH.",
    "6. Recreate the RFM segmentation (see /data/customer_rfm_segments.csv) using",
    "   nested IF() formulas instead of Python.",
]
for i, line in enumerate(lines, start=2):
    ws.cell(row=i, column=1, value=line).font = NORMAL_FONT
ws.column_dimensions["A"].width = 95

wb.save(OUT_PATH)
print(f"Saved workbook to {OUT_PATH}")
print(f"Months: {len(months)} | Products: {len(prod_names)} | Customers: {len(unique_customers)}")
